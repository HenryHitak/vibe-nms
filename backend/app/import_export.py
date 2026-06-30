from __future__ import annotations

import io
import json
import sqlite3
import zipfile
from typing import Any

from openpyxl import Workbook, load_workbook

from .audit import changed_fields, write_audit_log
from .db import DEVICE_COLUMNS, row_to_dict, rows_to_dicts
from .security import Actor
from .validation import (
    VALID_CRITICALITY,
    VALID_DEVICE_TYPES,
    ip_in_allowed_networks,
    normalize_upper,
    parse_bool,
    validate_ip,
    validate_mac,
)


TEMPLATE_COLUMNS = [
    ("Plant Code", "plant_code"),
    ("Plant Name", "plant_name"),
    ("Building", "building"),
    ("Floor", "floor"),
    ("Area", "area"),
    ("Zone", "zone"),
    ("Line Code", "line_code"),
    ("Line Name", "line_name"),
    ("Detailed Location", "detailed_location"),
    ("Device Name", "device_name"),
    ("Device Type", "device_type"),
    ("IP Address", "ip_address"),
    ("MAC Address", "mac_address"),
    ("Hostname", "hostname"),
    ("Connected AP Name", "connected_ap_name"),
    ("Connected AP IP", "connected_ap_ip"),
    ("Switch Name", "switch_name"),
    ("Switch Port", "switch_port"),
    ("VLAN", "vlan"),
    ("Owner Department", "owner_department"),
    ("Criticality", "criticality"),
    ("Monitoring Enabled", "monitoring_enabled"),
    ("Notes", "notes"),
]

HEADER_TO_FIELD = {header.lower(): field for header, field in TEMPLATE_COLUMNS}


def _cell_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, str):
        text = value.strip()
        return text if text else None
    return value


def _json(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False, default=str)


def build_template_workbook() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "devices"
    sheet.append([header for header, _field in TEMPLATE_COLUMNS])
    sample = {
        "plant_code": "MX01",
        "plant_name": "Main Plant",
        "building": "A",
        "floor": "1",
        "area": "Assembly",
        "zone": "A1",
        "line_code": "LINE-01",
        "line_name": "Assembly Line 1",
        "detailed_location": "Scanner station",
        "device_name": "LINE1_SCANNER_03",
        "device_type": "SCANNER",
        "ip_address": "10.10.1.55",
        "mac_address": "00:11:22:33:44:55",
        "hostname": "line1-scanner-03",
        "connected_ap_name": "AP_LINE1_A",
        "connected_ap_ip": "10.10.1.10",
        "switch_name": "SW_LINE1",
        "switch_port": "Gi1/0/12",
        "vlan": 110,
        "owner_department": "Production",
        "criticality": "HIGH",
        "monitoring_enabled": "TRUE",
        "notes": "Replace this sample row",
    }
    sheet.append([sample[field] for _header, field in TEMPLATE_COLUMNS])
    for column in sheet.columns:
        max_len = max(len(str(cell.value or "")) for cell in column)
        sheet.column_dimensions[column[0].column_letter].width = min(max(max_len + 2, 12), 32)
    return workbook_to_bytes(workbook)


def workbook_to_bytes(workbook: Workbook) -> bytes:
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def _excel_value(value: Any) -> Any:
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False, default=str)
    return value


def workbook_from_rows(sheet_name: str, columns: list[tuple[str, str]], rows: list[dict[str, Any]]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_name[:31]
    sheet.append([header for header, _field in columns])
    for row in rows:
        sheet.append([_excel_value(row.get(field)) for _header, field in columns])
    for column in sheet.columns:
        max_len = max(len(str(cell.value or "")) for cell in column)
        sheet.column_dimensions[column[0].column_letter].width = min(max(max_len + 2, 12), 48)
    return workbook_to_bytes(workbook)


def _normalize_excel_rows(payload: bytes) -> list[tuple[int, dict[str, Any]]]:
    workbook = load_workbook(io.BytesIO(payload), data_only=True)
    sheet = workbook.active
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_row:
        return []
    fields = [HEADER_TO_FIELD.get(str(header or "").strip().lower()) for header in header_row]
    rows: list[tuple[int, dict[str, Any]]] = []
    for row_index, row_values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        row = {
            field: _cell_value(row_values[index] if index < len(row_values) else None)
            for index, field in enumerate(fields)
            if field
        }
        if any(value not in (None, "") for value in row.values()):
            rows.append((row_index, row))
    return rows


def validate_import_rows(conn: sqlite3.Connection, payload: bytes) -> list[dict[str, Any]]:
    normalized_rows = _normalize_excel_rows(payload)
    existing_ips = {
        row["ip_address"]
        for row in conn.execute("SELECT ip_address FROM network_devices WHERE ip_address IS NOT NULL").fetchall()
    }
    existing_macs = {
        row["mac_address"]: row["ip_address"]
        for row in conn.execute("SELECT mac_address, ip_address FROM network_devices WHERE mac_address IS NOT NULL AND mac_address != ''").fetchall()
    }
    seen_ips: set[str] = set()
    seen_macs: set[str] = set()
    results: list[dict[str, Any]] = []

    for row_number, row in normalized_rows:
        errors: list[str] = []
        warnings: list[str] = []

        for field in ("plant_code", "line_code", "device_name", "device_type", "ip_address"):
            if not row.get(field):
                errors.append(f"Missing {field.replace('_', ' ')}")

        ip_address = str(row.get("ip_address") or "").strip()
        if ip_address:
            if not validate_ip(ip_address):
                errors.append("Invalid IP address")
            elif ip_address in seen_ips:
                errors.append("Duplicate IP address in import file")
            else:
                seen_ips.add(ip_address)
                if ip_address in existing_ips:
                    warnings.append("IP exists and will be updated on commit")
                if not ip_in_allowed_networks(ip_address):
                    warnings.append("IP is outside configured corporate networks; monitoring will be restricted")

        mac_address = str(row.get("mac_address") or "").strip()
        if mac_address:
            normalized_mac = mac_address.upper().replace("-", ":")
            row["mac_address"] = normalized_mac
            if not validate_mac(normalized_mac):
                errors.append("Invalid MAC address")
            elif normalized_mac in seen_macs:
                errors.append("Duplicate MAC address in import file")
            else:
                seen_macs.add(normalized_mac)
                if normalized_mac in existing_macs and existing_macs[normalized_mac] != ip_address:
                    warnings.append("MAC exists on another device")

        device_type = normalize_upper(row.get("device_type"))
        if device_type:
            row["device_type"] = device_type
            if device_type not in VALID_DEVICE_TYPES:
                errors.append("Invalid device type")

        criticality = normalize_upper(row.get("criticality") or "MEDIUM")
        row["criticality"] = criticality
        if criticality not in VALID_CRITICALITY:
            errors.append("Invalid criticality")

        monitoring_enabled = parse_bool(row.get("monitoring_enabled") if row.get("monitoring_enabled") is not None else True)
        if monitoring_enabled is None:
            errors.append("Invalid monitoring enabled value")
        else:
            row["monitoring_enabled"] = monitoring_enabled

        vlan = row.get("vlan")
        if vlan in (None, ""):
            row["vlan"] = None
        else:
            try:
                vlan_number = int(vlan)
                if vlan_number < 1 or vlan_number > 4094:
                    errors.append("Invalid VLAN value")
                else:
                    row["vlan"] = vlan_number
            except (TypeError, ValueError):
                errors.append("Invalid VLAN value")

        connected_ap_ip = str(row.get("connected_ap_ip") or "").strip()
        if connected_ap_ip and not validate_ip(connected_ap_ip):
            errors.append("Invalid connected AP IP")
        if connected_ap_ip and connected_ap_ip == ip_address:
            errors.append("Invalid AP relationship")
        if row.get("connected_ap_name") and not connected_ap_ip:
            warnings.append("Connected AP name has no AP IP")

        status = "ERROR" if errors else "WARNING" if warnings else "VALID"
        results.append(
            {
                "row_number": row_number,
                "row_data": row,
                "validation_status": status,
                "validation_message": "; ".join(errors + warnings),
            }
        )
    return results


def create_import_job(conn: sqlite3.Connection, file_name: str, actor: Actor, results: list[dict[str, Any]]) -> dict[str, Any]:
    totals = {
        "total_rows": len(results),
        "valid_rows": sum(1 for row in results if row["validation_status"] == "VALID"),
        "warning_rows": sum(1 for row in results if row["validation_status"] == "WARNING"),
        "error_rows": sum(1 for row in results if row["validation_status"] == "ERROR"),
    }
    cursor = conn.execute(
        """
        INSERT INTO import_jobs(file_name, uploaded_by, uploaded_from_ip, status, total_rows, valid_rows, warning_rows, error_rows)
        VALUES (?, ?, ?, 'PREVIEWED', ?, ?, ?, ?)
        """,
        (
            file_name,
            actor.username,
            actor.ip_address,
            totals["total_rows"],
            totals["valid_rows"],
            totals["warning_rows"],
            totals["error_rows"],
        ),
    )
    import_job_id = cursor.lastrowid
    for row in results:
        conn.execute(
            """
            INSERT INTO import_job_rows(import_job_id, row_number, row_data_json, validation_status, validation_message)
            VALUES (?, ?, ?, ?, ?)
            """,
            (
                import_job_id,
                row["row_number"],
                _json(row["row_data"]),
                row["validation_status"],
                row["validation_message"],
            ),
        )
    job = row_to_dict(conn.execute("SELECT * FROM import_jobs WHERE id = ?", (import_job_id,)).fetchone()) or {}
    job["rows"] = results[:200]
    return job


def commit_import_job(conn: sqlite3.Connection, import_job_id: int, actor: Actor) -> dict[str, Any]:
    job = conn.execute("SELECT * FROM import_jobs WHERE id = ?", (import_job_id,)).fetchone()
    if not job:
        raise ValueError("Import job not found")
    if job["status"] == "COMMITTED":
        raise ValueError("Import job already committed")

    rows = conn.execute(
        """
        SELECT * FROM import_job_rows
        WHERE import_job_id = ? AND validation_status IN ('VALID', 'WARNING')
        ORDER BY row_number
        """,
        (import_job_id,),
    ).fetchall()

    inserted = 0
    updated = 0
    failed = 0
    errors: list[str] = []

    for import_row in rows:
        row_data = json.loads(import_row["row_data_json"])
        try:
            device = {column: row_data.get(column) for column in DEVICE_COLUMNS}
            device["monitoring_enabled"] = 1 if device.get("monitoring_enabled") else 0
            existing = conn.execute(
                "SELECT * FROM network_devices WHERE ip_address = ?",
                (device["ip_address"],),
            ).fetchone()
            if existing:
                before = row_to_dict(existing) or {}
                assignments = ", ".join([f"{column} = ?" for column in DEVICE_COLUMNS])
                values = [device[column] for column in DEVICE_COLUMNS]
                conn.execute(
                    f"""
                    UPDATE network_devices
                    SET {assignments}, updated_by = ?, updated_from_ip = ?, updated_at = CURRENT_TIMESTAMP, is_deleted = 0,
                        deleted_by = NULL, deleted_from_ip = NULL, deleted_at = NULL
                    WHERE id = ?
                    """,
                    values + [actor.username, actor.ip_address, existing["id"]],
                )
                after = row_to_dict(conn.execute("SELECT * FROM network_devices WHERE id = ?", (existing["id"],)).fetchone()) or {}
                write_audit_log(
                    conn,
                    actor,
                    "UPDATE",
                    "DEVICE",
                    entity_id=existing["id"],
                    target_ip_address=device["ip_address"],
                    before_data=before,
                    after_data=after,
                    changed=changed_fields(before, after),
                )
                updated += 1
            else:
                columns = DEVICE_COLUMNS + ["created_by", "created_from_ip"]
                values = [device[column] for column in DEVICE_COLUMNS] + [actor.username, actor.ip_address]
                placeholders = ", ".join(["?"] * len(columns))
                cursor = conn.execute(
                    f"INSERT INTO network_devices({', '.join(columns)}) VALUES ({placeholders})",
                    values,
                )
                after = row_to_dict(conn.execute("SELECT * FROM network_devices WHERE id = ?", (cursor.lastrowid,)).fetchone()) or {}
                write_audit_log(
                    conn,
                    actor,
                    "CREATE",
                    "DEVICE",
                    entity_id=cursor.lastrowid,
                    target_ip_address=device["ip_address"],
                    before_data=None,
                    after_data=after,
                    changed={},
                )
                inserted += 1
        except Exception as exc:  # pragma: no cover - kept visible in import job result
            failed += 1
            errors.append(f"Row {import_row['row_number']}: {exc}")

    conn.execute(
        """
        UPDATE import_jobs
        SET status = 'COMMITTED', inserted_rows = ?, updated_rows = ?, failed_rows = ?, completed_at = CURRENT_TIMESTAMP
        WHERE id = ?
        """,
        (inserted, updated, failed, import_job_id),
    )
    summary = {
        "import_job_id": import_job_id,
        "inserted_rows": inserted,
        "updated_rows": updated,
        "failed_rows": failed,
        "errors": errors,
    }
    write_audit_log(
        conn,
        actor,
        "IMPORT",
        "DEVICE",
        entity_id=import_job_id,
        after_data=summary,
        changed={},
        result="SUCCESS" if failed == 0 else "FAILED",
        error_message="; ".join(errors[:5]) if errors else None,
    )
    return summary


def devices_rows(conn: sqlite3.Connection, include_deleted: bool = False) -> list[dict[str, Any]]:
    where = "" if include_deleted else "WHERE is_deleted = 0"
    return rows_to_dicts(conn.execute(f"SELECT * FROM network_devices {where} ORDER BY plant_code, line_code, device_name").fetchall())


def plants_rows(conn: sqlite3.Connection) -> list[dict[str, Any]]:
    return rows_to_dicts(
        conn.execute(
            """
            SELECT plant_code, MAX(plant_name) AS plant_name, COUNT(*) AS device_count
            FROM network_devices
            WHERE is_deleted = 0
            GROUP BY plant_code
            ORDER BY plant_code
            """
        ).fetchall()
    )


def access_points_rows(conn: sqlite3.Connection) -> list[dict[str, Any]]:
    return rows_to_dicts(
        conn.execute(
            """
            SELECT connected_ap_name, connected_ap_ip, plant_code, line_code, COUNT(*) AS connected_device_count
            FROM network_devices
            WHERE is_deleted = 0 AND connected_ap_name IS NOT NULL AND connected_ap_name != ''
            GROUP BY connected_ap_name, connected_ap_ip, plant_code, line_code
            ORDER BY plant_code, line_code, connected_ap_name
            """
        ).fetchall()
    )


def lines_rows(conn: sqlite3.Connection) -> list[dict[str, Any]]:
    return rows_to_dicts(
        conn.execute(
            """
            SELECT plant_code, line_code, MAX(line_name) AS line_name, COUNT(*) AS device_count
            FROM network_devices
            WHERE is_deleted = 0
            GROUP BY plant_code, line_code
            ORDER BY plant_code, line_code
            """
        ).fetchall()
    )


def locations_rows(conn: sqlite3.Connection) -> list[dict[str, Any]]:
    return rows_to_dicts(
        conn.execute(
            """
            SELECT plant_code, building, floor, area, zone, COUNT(*) AS device_count
            FROM network_devices
            WHERE is_deleted = 0
            GROUP BY plant_code, building, floor, area, zone
            ORDER BY plant_code, building, floor, area, zone
            """
        ).fetchall()
    )


def export_job(conn: sqlite3.Connection, actor: Actor, export_type: str, file_name: str, row_count: int) -> None:
    conn.execute(
        """
        INSERT INTO export_jobs(export_type, requested_by, requested_from_ip, file_name, row_count)
        VALUES (?, ?, ?, ?, ?)
        """,
        (export_type, actor.username, actor.ip_address, file_name, row_count),
    )
    write_audit_log(
        conn,
        actor,
        "EXPORT",
        "SYSTEM_SETTING" if export_type == "full-backup" else "DEVICE",
        entity_id=file_name,
        after_data={"export_type": export_type, "file_name": file_name, "row_count": row_count},
        changed={},
    )


def devices_workbook(conn: sqlite3.Connection, include_deleted: bool = True) -> bytes:
    rows = devices_rows(conn, include_deleted=include_deleted)
    columns = TEMPLATE_COLUMNS + [
        ("Status", "status"),
        ("Latency Ms", "latency_ms"),
        ("Packet Loss Percent", "packet_loss_percent"),
        ("Consecutive Failure Count", "consecutive_failure_count"),
        ("Is Deleted", "is_deleted"),
    ]
    return workbook_from_rows("devices", columns, rows)


def audit_logs_workbook(conn: sqlite3.Connection) -> bytes:
    rows = rows_to_dicts(conn.execute("SELECT * FROM audit_logs ORDER BY created_at DESC LIMIT 10000").fetchall())
    columns = [(key.replace("_", " ").title(), key) for key in (rows[0].keys() if rows else ["id", "created_at"])]
    return workbook_from_rows("audit_logs", columns, rows)


def simple_rows_workbook(sheet_name: str, rows: list[dict[str, Any]]) -> bytes:
    columns = [(key.replace("_", " ").title(), key) for key in (rows[0].keys() if rows else ["empty"])]
    return workbook_from_rows(sheet_name, columns, rows)


def migration_payload(conn: sqlite3.Connection) -> dict[str, Any]:
    tables = [
        "network_devices",
        "audit_logs",
        "import_jobs",
        "export_jobs",
        "monitoring_runs",
        "device_metrics",
        "alerts",
        "notifications",
        "system_settings",
    ]
    payload: dict[str, Any] = {"version": 1, "tables": {}}
    for table in tables:
        payload["tables"][table] = rows_to_dicts(conn.execute(f"SELECT * FROM {table}").fetchall())
    return payload


def full_backup_zip(conn: sqlite3.Connection) -> bytes:
    output = io.BytesIO()
    payload = migration_payload(conn)
    with zipfile.ZipFile(output, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        archive.writestr("devices.xlsx", devices_workbook(conn, include_deleted=True))
        archive.writestr("plants.xlsx", simple_rows_workbook("plants", plants_rows(conn)))
        archive.writestr("locations.xlsx", simple_rows_workbook("locations", locations_rows(conn)))
        archive.writestr("production_lines.xlsx", simple_rows_workbook("production_lines", lines_rows(conn)))
        archive.writestr("access_points.xlsx", simple_rows_workbook("access_points", access_points_rows(conn)))
        archive.writestr("alerts.xlsx", simple_rows_workbook("alerts", rows_to_dicts(conn.execute("SELECT * FROM alerts").fetchall())))
        archive.writestr("audit_logs.xlsx", audit_logs_workbook(conn))
        archive.writestr("nms_config.json", json.dumps({"source": "vibe-nms", "version": 1}, indent=2))
        archive.writestr("migration.json", json.dumps(payload, indent=2, ensure_ascii=False, default=str))
        if getattr(conn, "engine", "sqlite") == "sqlite":
            try:
                conn.execute("PRAGMA wal_checkpoint(FULL)")
            except sqlite3.OperationalError:
                pass
            db_path = conn.execute("PRAGMA database_list").fetchone()["file"]
            with open(db_path, "rb") as database_file:
                archive.writestr("database_backup.sqlite", database_file.read())
        else:
            schema_path = __import__("pathlib").Path(__file__).resolve().parent.parent / "mssql" / "schema.sql"
            archive.writestr("mssql_schema.sql", schema_path.read_text(encoding="utf-8"))
            archive.writestr(
                "mssql_backup_note.txt",
                "Use SQL Server backup tooling for physical .bak backups. This zip includes migration.json and mssql_schema.sql.",
            )
    return output.getvalue()
